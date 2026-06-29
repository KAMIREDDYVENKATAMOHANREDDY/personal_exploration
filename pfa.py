import os
import zipfile
import shutil
from openpyxl import load_workbook


# Paths
TMR_LIST_FILE = r"D:\Input\TMR_List.xlsx"
ZIP_FOLDER = r"D:\TMR_Zips"
MASTER_FILE = r"D:\Master\Master.xlsx"
TEMP_FOLDER = r"D:\Temp\Extract"
LOG_FILE = r"D:\Output\TMR_Process_Log.xlsx"


# Read TMR list
def read_tmr_list():
    wb = load_workbook(TMR_LIST_FILE)
    ws = wb.active

    tmrs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tmrs.append(row[0])

    # Sort descending by number
    tmrs.sort(key=lambda x: int(x.split("-")[1]), reverse=True)

    return tmrs


# Find ZIP
def find_zip(tmr):
    for f in os.listdir(ZIP_FOLDER):
        if f.startswith(tmr) and f.endswith(".zip"):
            return os.path.join(ZIP_FOLDER, f)

    return None


# Find BMM Excel
def find_bmm(folder):
    for root, dirs, files in os.walk(folder):
        for f in files:
            if f.startswith("BMM") and f.endswith(".xlsx"):
                return os.path.join(root, f)

    return None


# Process
log = []

master_wb = load_workbook(MASTER_FILE)
master_ws = master_wb.active


for tmr in read_tmr_list():

    zip_file = find_zip(tmr)

    if not zip_file:
        log.append([tmr, "ZIP Missing", "-", 0, "Skipped"])
        continue


    # Extract
    if os.path.exists(TEMP_FOLDER):
        shutil.rmtree(TEMP_FOLDER)

    os.makedirs(TEMP_FOLDER)

    with zipfile.ZipFile(zip_file) as z:
        z.extractall(TEMP_FOLDER)


    bmm_file = find_bmm(TEMP_FOLDER)

    if not bmm_file:
        log.append([tmr, "Available", "BMM Missing", 0, "Skipped"])
        continue


    # Read BMM data
    src_wb = load_workbook(bmm_file)
    src_ws = src_wb.worksheets[0]

    data = []

    for row in src_ws.iter_rows(min_row=5, values_only=True):
        if row[1] and str(row[1]).startswith("TC_"):
            data.append(row)


    if data:

        # Insert rows at row 15
        master_ws.insert_rows(15, amount=len(data))

        # Paste data
        for r, row_data in enumerate(data, start=15):
            for c, value in enumerate(row_data, start=1):
                master_ws.cell(r, c).value = value


    log.append(
        [tmr, "Available", "Found", len(data), "Added"]
    )


# Save master
master_wb.save(MASTER_FILE)


# Create log workbook
log_wb = load_workbook()
log_ws = log_wb.active

log_ws.append(
    ["TMR", "ZIP Status", "BMM Status", "Rows Added", "Result"]
)

for item in log:
    log_ws.append(item)

log_wb.save(LOG_FILE)


print("Completed")